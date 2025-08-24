from rest_framework import generics, status, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth.hashers import check_password
from django.db.models import Sum, F, DecimalField, Value, Q
from django.db.models.functions import Coalesce
import logging

from .models import WorkCategory, User, AuthToken, Project, Estimate, WorkType, Status, WorkPrice, Role, ProjectAssignment, EstimateItem
from .serializers import (
    WorkCategorySerializer, LoginSerializer, UserSerializer, ProjectSerializer, 
    EstimateListSerializer, WorkTypeSerializer, StatusSerializer, EstimateDetailSerializer, RoleSerializer,
    ProjectAssignmentSerializer, EstimateItemSerializer
)
from .permissions import IsManager, IsAuthenticatedCustom, CanAccessEstimate
from .security_decorators import ensure_estimate_access, audit_critical_action, log_data_change
import openpyxl
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side
from collections import defaultdict

# Настройка логгеров
security_logger = logging.getLogger('security')
audit_logger = logging.getLogger('audit')

# Пагинация для работ
class WorkTypePagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class WorkTypeImportView(APIView):
    permission_classes = [IsAuthenticatedCustom, IsManager]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.data.get('file')
        if not file_obj:
            return Response({"error": "Файл для импорта не найден."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file_obj)
            sheet = workbook.active

            created_count = 0
            updated_count = 0
            errors = []

            # Пропускаем заголовки
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                name, category_name, unit, cost_price, client_price = row

                # Пропускаем пустые строки
                if not name:
                    continue

                # Валидация данных
                if not all([category_name, unit, cost_price, client_price]):
                    errors.append(f"Строка {row_idx}: не все поля заполнены.")
                    continue
                
                try:
                    cost_price = float(cost_price)
                    client_price = float(client_price)
                except (ValueError, TypeError):
                    errors.append(f"Строка {row_idx}: цены должны быть числами.")
                    continue

                # Получаем или создаем категорию
                category, _ = WorkCategory.objects.get_or_create(category_name=category_name)

                # Обновляем или создаем работу и ее цену
                work_type, created = WorkType.objects.update_or_create(
                    work_name=name,
                    defaults={
                        'category': category,
                        'unit_of_measurement': unit
                    }
                )

                WorkPrice.objects.update_or_create(
                    work_type=work_type,
                    defaults={
                        'cost_price': cost_price,
                        'client_price': client_price
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            response_data = {
                "message": "Импорт успешно завершен.",
                "created": created_count,
                "updated": updated_count,
                "errors": errors
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Произошла ошибка при обработке файла: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    permission_classes = []
    authentication_classes = []
    def post(self, request, *args, **kwargs):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            try:
                user = User.objects.select_related('role').get(email=email)
            except User.DoesNotExist:
                return Response({"error": "Invalid credentials"}, status=status.HTTP_400_BAD_REQUEST)

            if check_password(password, user.password_hash):
                token, _ = AuthToken.objects.update_or_create(user=user)
                return Response({
                    'token': str(token.token),
                    'user': UserSerializer(user).data
                })
            else:
                return Response({"error": "Invalid credentials"}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CurrentUserView(APIView):
    permission_classes = [IsAuthenticatedCustom]
    
    def get(self, request, *args, **kwargs):
        """Получить информацию о текущем пользователе"""
        return Response(UserSerializer(request.user).data)


# --- ViewSets для управления ---

class WorkCategoryViewSet(viewsets.ModelViewSet):
    queryset = WorkCategory.objects.all()
    serializer_class = WorkCategorySerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [IsAuthenticatedCustom]
        else:
            self.permission_classes = [IsAuthenticatedCustom, IsManager]
        return super().get_permissions()
    
    def perform_destroy(self, instance):
        """Переопределяем удаление для лучшего контроля ошибок"""
        try:
            # Проверяем, есть ли связанные работы
            related_works_count = WorkType.objects.filter(category=instance).count()
            if related_works_count > 0:
                from rest_framework.exceptions import ValidationError
                raise ValidationError(
                    f"Нельзя удалить категорию '{instance.category_name}'. "
                    f"С ней связано работ: {related_works_count}. "
                    f"Сначала удалите или перенесите все работы из этой категории."
                )
            instance.delete()
        except Exception as e:
            from rest_framework.exceptions import ValidationError
            if 'RESTRICT' in str(e) or 'foreign key constraint fails' in str(e):
                raise ValidationError(
                    f"Нельзя удалить категорию '{instance.category_name}'. "
                    f"Сначала удалите все связанные с ней работы."
                )
            raise ValidationError(f"Ошибка при удалении: {str(e)}")

class WorkTypeViewSet(viewsets.ModelViewSet):
    queryset = WorkType.objects.select_related('category', 'workprice').order_by('-usage_count')
    serializer_class = WorkTypeSerializer
    pagination_class = WorkTypePagination

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [IsAuthenticatedCustom]
        else:
            self.permission_classes = [IsAuthenticatedCustom, IsManager]
        return super().get_permissions()
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Если запрос содержит параметр 'all', отключаем пагинацию и меняем сортировку
        if self.request.query_params.get('all') == 'true':
            self.pagination_class = None
            # Для режима "все работы" сортируем по категории и названию, а не по usage_count
            queryset = WorkType.objects.select_related('category', 'workprice').order_by('category__category_name', 'work_name')
        return queryset

    def perform_create(self, serializer):
        # Извлекаем данные о ценах из запроса
        cost_price = serializer.validated_data.pop('cost_price')
        client_price = serializer.validated_data.pop('client_price')
        
        # Создаем WorkType
        work_type = serializer.save()
        
        # Создаем соответствующую запись в WorkPrice с указанными ценами
        WorkPrice.objects.create(
            work_type=work_type, 
            cost_price=cost_price, 
            client_price=client_price
        )
    
    def perform_update(self, serializer):
        # Проверяем, есть ли данные о ценах в запросе
        cost_price = serializer.validated_data.pop('cost_price', None)
        client_price = serializer.validated_data.pop('client_price', None)
        
        # Обновляем WorkType
        work_type = serializer.save()
        
        # Обновляем цены, если они переданы
        if cost_price is not None or client_price is not None:
            work_price, created = WorkPrice.objects.get_or_create(work_type=work_type)
            if cost_price is not None:
                work_price.cost_price = cost_price
            if client_price is not None:
                work_price.client_price = client_price
            work_price.save()
    
    def perform_destroy(self, instance):
        """Переопределяем удаление для лучшего контроля ошибок"""
        try:
            # Проверяем, используется ли эта работа в сметах
            from .models import EstimateItem
            related_items_count = EstimateItem.objects.filter(work_type=instance).count()
            if related_items_count > 0:
                from rest_framework.exceptions import ValidationError
                raise ValidationError(
                    f"Нельзя удалить работу '{instance.work_name}'. "
                    f"Она используется в {related_items_count} позициях смет. "
                    f"Сначала удалите эти позиции из смет."
                )
            instance.delete()
        except Exception as e:
            from rest_framework.exceptions import ValidationError
            if 'RESTRICT' in str(e) or 'foreign key constraint fails' in str(e):
                raise ValidationError(
                    f"Нельзя удалить работу '{instance.work_name}'. "
                    f"Она используется в существующих сметах."
                )
            raise ValidationError(f"Ошибка при удалении: {str(e)}")

class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role.role_name == 'менеджер':
            return Project.objects.all()
        else:
            return Project.objects.filter(projectassignment__user=user)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsAuthenticatedCustom, IsManager]
        else:
            self.permission_classes = [IsAuthenticatedCustom]
        return super().get_permissions()

class HealthCheckView(APIView):
    """
    Публичный endpoint для проверки работоспособности приложения
    """
    permission_classes = []  # Без аутентификации
    
    def get(self, request):
        return Response({
            'status': 'healthy',
            'service': 'estimate-backend',
            'version': '1.0.0'
        })

class StatusListView(generics.ListAPIView):
    queryset = Status.objects.all()
    serializer_class = StatusSerializer
    permission_classes = [IsAuthenticatedCustom]

class EstimateViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticatedCustom, CanAccessEstimate]

    def get_serializer_class(self):
        if self.action == 'list':
            return EstimateListSerializer
        return EstimateDetailSerializer

    def perform_create(self, serializer):
        user = self.request.user
        try:
            draft_status = Status.objects.get(status_name='Черновик')
        except Status.DoesNotExist:
            draft_status = None

        # ИСПРАВЛЕНИЕ: Получаем foreman_id из данных запроса
        foreman_id = self.request.data.get('foreman_id')
        
        if foreman_id:
            # Если передан foreman_id, используем его (менеджер назначает прораба)
            try:
                foreman = User.objects.get(user_id=foreman_id)
                audit_logger.info(
                    f"НАЗНАЧЕНИЕ ПРОРАБА: Менеджер {user.email} назначил прораба {foreman.email} на новую смету"
                )
            except User.DoesNotExist:
                foreman = user  # Fallback на текущего пользователя
                security_logger.warning(
                    f"FALLBACK: Некорректный foreman_id {foreman_id}, используется создатель {user.email}"
                )
        else:
            # Если foreman_id не передан, используем текущего пользователя
            foreman = user

        serializer.save(
            creator=user,
            foreman=foreman, 
            status=draft_status
        )

    def get_queryset(self):
        user = self.request.user
        
        # Базовый queryset с оптимизацией для связанных полей
        queryset = Estimate.objects.select_related(
            'project', 'creator', 'status', 'foreman'
        ).all()

        # КРИТИЧЕСКИ ВАЖНО: Фильтруем по роли пользователя для ВСЕХ операций
        if user.role.role_name != 'менеджер':
            # Прораб видит ТОЛЬКО те сметы, где он назначен прорабом
            queryset = queryset.filter(foreman=user)

        # Если это запрос на список, добавляем аннотацию с общей суммой
        if self.action == 'list':
            if user.role.role_name != 'менеджер':
                # ДЛЯ ПРОРАБОВ: всегда считаем только работы добавленные ими или без автора (старые)
                # Убираем различие между desktop и mobile - прораб везде видит только свои работы
                queryset = queryset.annotate(
                    totalAmount=Coalesce(
                        Sum(
                            F('items__quantity') * F('items__cost_price_per_unit'),
                            output_field=DecimalField(),
                            filter=Q(items__added_by=user) | Q(items__added_by__isnull=True)
                        ),
                        Value(0.0), # Если нет работ, вернуть 0.0
                        output_field=DecimalField()
                    ),
                    mobile_total_amount=Coalesce(
                        Sum(
                            F('items__quantity') * F('items__cost_price_per_unit'),
                            output_field=DecimalField(),
                            filter=Q(items__added_by=user) | Q(items__added_by__isnull=True)
                        ),
                        Value(0.0), # Если нет работ, вернуть 0.0
                        output_field=DecimalField()
                    )
                )
            else:
                # ДЛЯ МЕНЕДЖЕРОВ: полная сумма всех работ
                queryset = queryset.annotate(
                    totalAmount=Coalesce(
                        Sum(
                            F('items__quantity') * F('items__cost_price_per_unit'),
                            output_field=DecimalField()
                        ),
                        Value(0.0), # Если нет работ, вернуть 0.0
                        output_field=DecimalField()
                    ),
                    mobile_total_amount=Coalesce(
                        Sum(
                            F('items__quantity') * F('items__cost_price_per_unit'),
                            output_field=DecimalField()
                        ),
                        Value(0.0), # Если нет работ, вернуть 0.0
                        output_field=DecimalField()
                    )
                )
        else:
            # Для детального просмотра предзагружаем работы
            queryset = queryset.prefetch_related('items', 'items__work_type')
        
        return queryset

    def retrieve(self, request, *args, **kwargs):
        """Переопределяем retrieve для дополнительной проверки доступа и фильтрации работ"""
        instance = self.get_object()
        
        # КРИТИЧЕСКИ ВАЖНО: Дублируем проверку доступа для надежности
        if request.user.role.role_name != 'менеджер':
            if instance.foreman != request.user:
                security_logger.warning(
                    f"БЛОКИРОВАН ДОСТУП: Пользователь {request.user.email} пытался получить доступ к смете {instance.estimate_id}, "
                    f"но смета принадлежит {instance.foreman.email}"
                )
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Нет доступа к данной смете")
        
        audit_logger.info(f"ДОСТУП К СМЕТЕ: Пользователь {request.user.email} получил доступ к смете {instance.estimate_id}")
        
        # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: Фильтруем работы сметы по роли пользователя
        import logging
        logger = logging.getLogger('django')
        logger.warning(f"🔍 DEBUG retrieve: Пользователь {request.user.email}, роль: {request.user.role.role_name}")
        logger.warning(f"🔍 DEBUG retrieve: Смета {instance.estimate_id}, всего работ: {instance.items.count()}")
        
        if request.user.role.role_name != 'менеджер':
            # Для прорабов - показываем только их работы (added_by = текущий пользователь или NULL для старых работ)
            filtered_items = instance.items.filter(
                Q(added_by=request.user) | Q(added_by__isnull=True)
            )
            # Временно заменяем items на отфильтрованные
            instance._filtered_items = list(filtered_items)
            logger.warning(f"🔍 DEBUG retrieve: Прораб - отфильтровано работ: {len(instance._filtered_items)}")
            for item in instance._filtered_items:
                logger.warning(f"  - Работа: {item.work_type.work_name}, added_by: {item.added_by_id}")
        else:
            # Для менеджеров - все работы
            instance._filtered_items = list(instance.items.all())
            logger.warning(f"🔍 DEBUG retrieve: Менеджер - показываем все работы: {len(instance._filtered_items)}")
        
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        """Переопределяем update для дополнительной проверки доступа"""
        instance = self.get_object()
        
        # Дублируем проверку доступа для надежности
        if request.user.role.role_name != 'менеджер':
            if instance.foreman != request.user:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Нет доступа к данной смете")
        
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Переопределяем destroy для дополнительной проверки доступа"""
        instance = self.get_object()
        
        # Дублируем проверку доступа для надежности
        if request.user.role.role_name != 'менеджер':
            if instance.foreman != request.user:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Нет доступа к данной смете")
        
        return super().destroy(request, *args, **kwargs)

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related('role').all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticatedCustom, IsManager]

class RoleViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticatedCustom, IsManager]

class ProjectAssignmentViewSet(viewsets.ModelViewSet):
    queryset = ProjectAssignment.objects.select_related('project', 'user').all()
    serializer_class = ProjectAssignmentSerializer
    permission_classes = [IsAuthenticatedCustom, IsManager]


class EstimateExportBaseView(APIView):
    """Базовый класс для экспорта смет в Excel"""
    permission_classes = [IsAuthenticatedCustom]

    def get_estimate(self, estimate_id):
        """Получить смету с проверкой доступа"""
        try:
            estimate = Estimate.objects.select_related(
                'project', 'creator', 'status', 'foreman'
            ).prefetch_related('items', 'items__work_type').get(estimate_id=estimate_id)
            
            # Проверяем права доступа
            user = self.request.user
            if user.role.role_name != 'менеджер':
                # Прораб может экспортировать только свои сметы
                if estimate.foreman != user:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("Нет доступа к данной смете")
            
            return estimate
        except Estimate.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound("Смета не найдена")

    def create_excel_file(self, estimate, include_cost_prices=True, is_client_export=False):
        """Создать Excel файл со сметой"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Смета"

        # Стили
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        if include_cost_prices and not is_client_export:
            # Внутренний экспорт - новый формат
            # B1: Название сметы
            ws.cell(row=current_row, column=2, value=estimate.estimate_number or 'Без названия').font = title_font
            
            # C1: "Составил:", D1: имя пользователя  
            ws.cell(row=current_row, column=3, value="Составил:")
            ws.cell(row=current_row, column=4, value=estimate.creator.full_name if estimate.creator else 'Не указан')
            
            # E1: "Дата создания:", F1: дата создания
            ws.cell(row=current_row, column=5, value="Дата создания:")
            created_date = estimate.created_at.strftime('%d.%m.%Y') if estimate.created_at else 'Не указана'
            ws.cell(row=current_row, column=6, value=created_date)
            
            current_row += 1
            current_row += 1  # Пустая строка
            
            # A3:I3 - объединенная ячейка с названием, объектом и адресом
            ws.merge_cells('A3:I3')
            project_address = getattr(estimate.project, 'address', '') or 'Адрес не указан'
            merged_text = f"{estimate.estimate_number or 'Без названия'} {estimate.project.project_name} {project_address}"
            merged_cell = ws.cell(row=3, column=1, value=merged_text)
            merged_cell.font = Font(bold=True, size=12)
            merged_cell.alignment = Alignment(horizontal='center')
            
            current_row = 4
            current_row += 1  # Пустая строка после объединенной ячейки
        else:
            # Клиентский экспорт - убираем служебные данные
            current_row += 1  # Пустая строка
            current_row += 1  # Еще одна пустая строка
            
            # A3:F3 - объединенная ячейка с названием, адресом и датой
            ws.merge_cells('A3:F3')
            project_address = getattr(estimate.project, 'address', '') or 'Адрес не указан'
            created_date = estimate.created_at.strftime('%d.%m.%Y') if estimate.created_at else 'Дата не указана'
            merged_text = f"{estimate.estimate_number or 'Без названия'} {project_address} от {created_date}"
            merged_cell = ws.cell(row=3, column=1, value=merged_text)
            merged_cell.font = Font(bold=True, size=12)
            merged_cell.alignment = Alignment(horizontal='center')
            
            current_row = 4
            current_row += 1  # Пустая строка после объединенной ячейки

        # Заголовки таблицы
        headers = ["№", "Наименование работ", "Ед. изм.", "Кол-во"]
        
        if include_cost_prices:
            headers.extend(["Цена себестоимости", "Сумма себестоимости", "Цена клиента", "Сумма клиента", "Прибыль"])
        else:
            headers.extend(["Цена за ед.", "Общая сумма"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        current_row += 1

        # Данные таблицы, сгруппированные по категориям
        items = estimate.items.select_related('work_type__category').all()
        
        # Группируем по категориям
        categories_dict = defaultdict(list)
        for item in items:
            category_name = item.work_type.category.category_name if item.work_type.category else 'Без категории'
            categories_dict[category_name].append(item)
        
        total_cost = 0
        total_client = 0
        total_profit = 0
        item_counter = 1
        
        # Обработка каждой категории
        for category_name, category_items in categories_dict.items():
            # Заголовок категории
            category_cell = ws.cell(row=current_row, column=2, value=category_name.upper())
            category_cell.font = Font(bold=True, size=12)
            current_row += 1
            
            # Переменные для итогов по категории
            category_cost = 0
            category_client = 0
            category_profit = 0
            
            # Обработка работ в категории
            for item in category_items:
                ws.cell(row=current_row, column=1, value=item_counter).border = border
                ws.cell(row=current_row, column=2, value=item.work_type.work_name).border = border
                ws.cell(row=current_row, column=3, value=item.work_type.unit_of_measurement).border = border
                ws.cell(row=current_row, column=4, value=float(item.quantity)).border = border

                cost_total = float(item.quantity) * float(item.cost_price_per_unit)
                client_total = float(item.quantity) * float(item.client_price_per_unit)
                profit_total = client_total - cost_total
                
                category_cost += cost_total
                category_client += client_total
                category_profit += profit_total
                
                total_cost += cost_total
                total_client += client_total
                total_profit += profit_total

                if include_cost_prices:
                    ws.cell(row=current_row, column=5, value=float(item.cost_price_per_unit)).border = border
                    ws.cell(row=current_row, column=6, value=cost_total).border = border
                    ws.cell(row=current_row, column=7, value=float(item.client_price_per_unit)).border = border
                    ws.cell(row=current_row, column=8, value=client_total).border = border
                    ws.cell(row=current_row, column=9, value=profit_total).border = border
                else:
                    ws.cell(row=current_row, column=5, value=float(item.client_price_per_unit)).border = border
                    ws.cell(row=current_row, column=6, value=client_total).border = border

                current_row += 1
                item_counter += 1
            
            # Итог по категории
            if include_cost_prices:
                ws.cell(row=current_row, column=2, value="Итого по разделу:").font = Font(bold=True)
                ws.cell(row=current_row, column=6, value=category_cost).font = Font(bold=True)
                ws.cell(row=current_row, column=8, value=category_client).font = Font(bold=True)
                ws.cell(row=current_row, column=9, value=category_profit).font = Font(bold=True)
            else:
                ws.cell(row=current_row, column=2, value="Итого по разделу:").font = Font(bold=True)
                ws.cell(row=current_row, column=6, value=category_client).font = Font(bold=True)
            
            current_row += 2  # Добавляем пустую строку между категориями

        # Общий итог
        current_row += 1  # Еще одна пустая строка перед итогом
        if include_cost_prices:
            ws.cell(row=current_row, column=2, value="ОБЩИЙ ИТОГ:").font = Font(bold=True, size=14)
            ws.cell(row=current_row, column=6, value=total_cost).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=8, value=total_client).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=9, value=total_profit).font = Font(bold=True, size=12)
        else:
            ws.cell(row=current_row, column=2, value="ОБЩАЯ СУММА:").font = Font(bold=True, size=14)
            ws.cell(row=current_row, column=6, value=total_client).font = Font(bold=True, size=12)

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        return wb


class EstimateClientExportView(EstimateExportBaseView):
    """Экспорт сметы для клиента (без себестоимости)"""
    
    def get(self, request, estimate_id):
        estimate = self.get_estimate(estimate_id)
        wb = self.create_excel_file(estimate, include_cost_prices=False, is_client_export=True)
        
        # Создаем HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{estimate.estimate_number or 'Смета'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class EstimateInternalExportView(EstimateExportBaseView):
    """Внутренний экспорт сметы (с полными данными)"""
    
    def get(self, request, estimate_id):
        estimate = self.get_estimate(estimate_id)
        wb = self.create_excel_file(estimate, include_cost_prices=True)
        
        # Создаем HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"ВН_{estimate.estimate_number or 'Смета'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class EstimateItemViewSet(viewsets.ModelViewSet):
    """ViewSet для работы с элементами смет"""
    serializer_class = EstimateItemSerializer
    
    def get_permissions(self):
        """
        Устанавливает permissions в зависимости от действия
        """
        if self.action in ['update', 'partial_update', 'destroy']:
            # Для редактирования и удаления нужны дополнительные проверки авторства
            from .permissions import IsAuthenticatedCustom, CanEditEstimateItem
            self.permission_classes = [IsAuthenticatedCustom, CanEditEstimateItem]
        else:
            # Для остальных действий базовая аутентификация
            from .permissions import IsAuthenticatedCustom
            self.permission_classes = [IsAuthenticatedCustom]
        return super().get_permissions()
    
    def get_queryset(self):
        """
        КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ:
        ПРОРАБЫ везде (десктоп И мобильный) видят только свои работы!
        Менеджеры видят все работы.
        """
        user = self.request.user
        estimate_id = self.request.query_params.get('estimate')
        
        if estimate_id:
            queryset = EstimateItem.objects.filter(estimate_id=estimate_id)
            
            if user.role.role_name != 'менеджер':
                # Проверяем доступ к смете - прораб может работать только со своими сметами
                queryset = queryset.filter(estimate__foreman=user)
                
                # И видит только свои работы (добавленные им или старые без автора)
                queryset = queryset.filter(
                    Q(added_by=user) | Q(added_by__isnull=True)
                )
                
            return queryset
        
        # Если не указана конкретная смета
        if user.role.role_name == 'менеджер':
            return EstimateItem.objects.all()
        else:
            # Прораб видит только свои работы в своих сметах
            return EstimateItem.objects.filter(
                estimate__foreman=user
            ).filter(
                Q(added_by=user) | Q(added_by__isnull=True)
            )
    
    def perform_create(self, serializer):
        # Дополнительная проверка доступа при создании
        estimate = serializer.validated_data.get('estimate')
        user = self.request.user
        
        if user.role.role_name != 'менеджер':
            # Прораб может добавлять элементы только в свои сметы
            if estimate.foreman != user:
                raise PermissionError("Нет доступа к данной смете")
        
        # НОВАЯ ЛОГИКА: Автоматически устанавливаем added_by при создании
        serializer.save(added_by=user)
